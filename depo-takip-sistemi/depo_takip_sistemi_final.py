"""
Python Version: 3.12
File name: depo_takip_sistemi_final.py
Created Date: 2024-03-27
Author: Cemil İlkim Teke
Pylint Score: 8.36/10
"""
import os
import random
import sys
from sys import exit
import socket
import time
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.mime.text import MIMEText
import hashlib
from playsound import playsound
import winsound
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
import colorama
from colorama import Fore
from veri_tabani import *

def safe_chdir(path):
    try:
        os.chdir(path)
    except UnicodeEncodeError:
        encoded_path = path.encode('utf-8').decode('utf-8')
        os.chdir(encoded_path)

if getattr(sys, 'frozen', False):
    safe_chdir(os.path.dirname(sys.executable))
else:
    safe_chdir(os.path.dirname(os.path.abspath(__file__)))
# .exe dosyasının çalışma dizinini bulunduğu klasör olarak belirliyor
cwd = os.getcwd()
# print(cwd) (Çalışma Dizini Kontrolü)
try:
    bilgisayar_adi = socket.gethostname().capitalize()
except Exception:
    bilgisayar_adi = "Kullanıcı Adı Alınamadı"
try:
    ip_adresi = socket.gethostbyname(bilgisayar_adi)
except Exception:
    ip_adresi = "İp Adresi Alınamadı"
# Bilgisayardaki kullanıcı adını ve ip adresini okur
colorama.init(autoreset=True)
# Yazı Stillerini Birbirinden Ayrı Tutuyor
simdi = datetime.now()
tarih = simdi.strftime('%d.%m.%Y %H:%M')

# colorama modülündeki renklerden rastgele parlak bir renk seçer
renkler = [Fore.LIGHTRED_EX, Fore.LIGHTWHITE_EX, Fore.LIGHTGREEN_EX, Fore.LIGHTYELLOW_EX, Fore.LIGHTCYAN_EX, Fore.LIGHTBLUE_EX, Fore.LIGHTMAGENTA_EX]
rastgele_renk = random.choice(renkler)

def mail_gonder():
    try:
        gonderen_mail = "depotakipsistemi@gmail.com"
        mail = ""
        while True:
            print(f"""
{Fore.LIGHTWHITE_EX}Lütfen Seçiminizi Yapınız:
{Fore.CYAN}1{Fore.RESET} - Hızlı Erişimlerden Seç
{Fore.CYAN}2{Fore.RESET} - Maili Klavyeden Gir
""")
            mail_secim = int(input(""))
            if mail_secim == 1:
                hizli_erisim_goster()
                mail = hizli_erisim_sec()
                if mail is None:
                    continue  # Döngüyü tekrar başlat
                break  # Döngüyü kır
            elif mail_secim == 2:
                mail = str(input("\nLütfen Alıcı Email Adresini Giriniz\n"))
                break  # Döngüyü kır
            else:
                print(f"\n{Fore.LIGHTRED_EX}Lütfen Geçerli Bir Sayı Giriniz! (1-2)\n")
                playsound(f"{cwd}//assests//error_sound.wav", block=False)
                continue  # Döngüyü tekrar başlat
        print(f"{Fore.LIGHTGREEN_EX}Sunucuya Bağlanılıyor...\n{Fore.RESET}")
        subject = "Depo Takip Sistemi"
        body = "Tarafınıza Gönderilen Depo Takip Sistemi Dosyaları Ektedir. Lütfen Bu Maile Cevap Vermeyiniz!"
        # Mail bağlantısı oluşturma
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(gonderen_mail, "zqbkgmzlhhfuuowu")
        # Mail'i oluşturur
        msg = MIMEMultipart()
        msg['From'] = gonderen_mail
        msg['To'] = mail
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))
        # Dosya eklemek için dosya yolunu belirtir
        attachment = ""
        dosya_adi = ""
        print(f"""{Fore.LIGHTWHITE_EX}
Lütfen Mail Göndermek İstediğiniz Dosyayı Seçiniz:
{Fore.CYAN}1{Fore.RESET}- Kayıtlar
{Fore.CYAN}2{Fore.RESET}- Loglar
{Fore.CYAN}3{Fore.RESET}- İkisini Birden Gönder\n""")
        dosya_secim = int(input())
        if dosya_secim == 1:
            dosya_adi = "Kayıtlar.xlsx"
            attachment = open(f"{cwd}//Kayıtlar.xlsx", "rb")
        elif dosya_secim == 2:
            dosya_adi = "Loglar.txt"
            attachment = open(f"{cwd}//Loglar.txt", "rb")
        elif dosya_secim == 3:
            dosya_adi = "Loglar.txt Kayıtlar.xlsx"
            attachment = [open("Loglar.txt", "rb"), open("Kayıtlar.xlsx", "rb")]
        else:
            print(f"\n{Fore.LIGHTRED_EX}Lütfen Geçerli Bir Sayı Giriniz! (1-3)\n")
            playsound(f"{cwd}//assests//error_sound.wav", block=False)
            mail_gonder()
        # Dosya ekler
        if isinstance(attachment, list):
            for att in attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(att.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', "attachment", filename=att.name)
                msg.attach(part)
        else:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', "attachment", filename=dosya_adi)
            msg.attach(part)
        # E-postayı gönderir
        server.sendmail(gonderen_mail, mail, msg.as_string())
        print(f"\n{Fore.LIGHTGREEN_EX}Mail Başarıyla Gönderildi")
        # Bağlantıyı kapatır
        server.quit()
    except TimeoutError:
        print(f"\n{Fore.LIGHTRED_EX}Mail Gönderilirken Zaman Aşımına Uğradı!\nSMTP Portunun Doğru Olduğuna Emin Olun!")
        playsound(f"{cwd}//assests//error_sound.wav", block=False)
    except smtplib.SMTPRecipientsRefused:
        print(
            f"\n{Fore.LIGHTRED_EX}Mail Gönderilirken Bir Hata Ortaya Çıktı!\nAlıcı Adresini Doğru Yazdığınızdan Emin Olun!")
        playsound(f"{cwd}//assests//error_sound.wav", block=False)
    except smtplib.SMTPResponseException:
        print(
            f"\n{Fore.LIGHTRED_EX}Mail Gönderilirken Bir Hata Ortaya Çıktı!\nGönderici Adresini Doğru Yazdığınızdan Emin Olun!")
        playsound(f"{cwd}//assests//error_sound.wav", block=False)
    except socket.gaierror:
        print(f"\n{Fore.LIGHTRED_EX}Sunucuya Bağlanılamadı!\nİnternete Bağlı Olduğunuzdan Emin Olun!")
        playsound(f"{cwd}//assests//error_sound.wav", block=False)
    except ValueError:
        print(f"\n{Fore.LIGHTRED_EX}Lütfen Bir Sayı Giriniz!")
        playsound(f"{cwd}//assests//error_sound.wav", block=False)
        return mail_gonder()
    except FileNotFoundError:
        print(f"\n{Fore.LIGHTRED_EX}Excel Dosyası Bulunamadı. Yeni Bir Tane Oluşturuluyor...")
        playsound(f"{cwd}//assests//error_sound.wav", block=False)
        time.sleep(1.13)
        tablo_olustur()
    except Exception as error:
        print(f"{Fore.LIGHTRED_EX}Beklenmeyen Bir Hata Oluştu!\nHata Kodunu Yazdırmak İster Misiniz?\n")
        hata_secim = int(input(f"\n{Fore.CYAN}1{Fore.RESET} - Evet\n{Fore.CYAN}2{Fore.RESET} - Hayır\n"))
        if hata_secim == 1:
            with open("Hatalar.txt", "a", encoding='utf-8') as error_file:
                error_file.write(f"{tarih} Tarihinde Bu Hata Alınmıştır:\n{error}\n\n")
            print(f"{Fore.LIGHTRED_EX}Sistemden Çıkış Yapılıyor...")
            playsound(f"{cwd}//assests//shutdown_sound.wav", block=False)
            time.sleep(1.13)
            exit()
        else:
            print(f"{Fore.LIGHTRED_EX}Sistemden Çıkış Yapılıyor...")
            playsound(f"{cwd}//assests//shutdown_sound.wav", block=False)
            time.sleep(1.13)
            exit()


def tablo_olustur():
    try:
        wb = Workbook()
        tekirdag = wb.create_sheet(title="Tekirdağ")
        izmir = wb.create_sheet(title="İzmir")
        canakkale = wb.create_sheet(title="Çanakkale")
        # Excel Kitabında Sayfaları Oluşturur
        tekirdag.append(["Barkod Numarası", "Ürün Adı", "Miktar"])
        izmir.append(["Barkod Numarası", "Ürün Adı", "Miktar"])
        canakkale.append(["Barkod Numarası", "Ürün Adı", "Miktar"])
        # Sayfalara Header'ları yazar
        kenarlik = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                          bottom=Side(style='thin'))
        for i in ["A", "B", "C"]:
            tekirdag[f'{i}1'].alignment = Alignment(horizontal='center')
            tekirdag[f'{i}1'].border = kenarlik
            tekirdag[f'{i}1'].fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            tekirdag.column_dimensions['A'].width = len("Barkod Numarası") + 2
            tekirdag.column_dimensions['B'].width = len("Ürün Adı") + 2
            tekirdag.column_dimensions['C'].width = len("Miktar") + 2
            izmir[f'{i}1'].alignment = Alignment(horizontal='center')
            izmir[f'{i}1'].border = kenarlik
            izmir[f'{i}1'].fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            izmir.column_dimensions['A'].width = len("Barkod Numarası") + 2
            izmir.column_dimensions['B'].width = len("Ürün Adı") + 2
            izmir.column_dimensions['C'].width = len("Miktar") + 2
            canakkale[f'{i}1'].alignment = Alignment(horizontal='center')
            canakkale[f'{i}1'].border = kenarlik
            canakkale[f'{i}1'].fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            canakkale.column_dimensions['A'].width = len("Barkod Numarası") + 2
            canakkale.column_dimensions['B'].width = len("Ürün Adı") + 2
            canakkale.column_dimensions['C'].width = len("Miktar") + 2
            # Hücre Genişliğini Ayarlar
        sayfa_adi = wb.sheetnames[0]
        ilk_sayfa = wb[sayfa_adi]
        wb.remove(ilk_sayfa)
        # Otamatik olarak oluşturulan "Sheet" sayfasını siler
        wb.save(f"{cwd}//Kayıtlar.xlsx")
    except Exception as error:
        print(f"{Fore.LIGHTRED_EX}Beklenmeyen Bir Hata Oluştu!\nHata Kodunu Yazdırmak İster Misiniz?\n")
        hata_secim = int(input(f"\n{Fore.CYAN}1{Fore.RESET} - Evet\n{Fore.CYAN}2{Fore.RESET} - Hayır\n"))
        if hata_secim == 1:
            with open("Hatalar.txt", "a", encoding='utf-8') as error_file:
                error_file.write(f"{tarih} Tarihinde Bu Hata Alınmıştır:\n{error}\n\n")
            print(f"{Fore.LIGHTRED_EX}Sistemden Çıkış Yapılıyor...")
            playsound(f"{cwd}//assests//shutdown_sound.wav", block=False)
            time.sleep(1.13)
            exit()

        else:
            print(f"{Fore.LIGHTRED_EX}Sistemden Çıkış Yapılıyor...")
            playsound(f"{cwd}//assests//shutdown_sound.wav", block=False)
            time.sleep(1.13)
            exit()


def urun_ekle():
    try:
        wb = load_workbook(f"{cwd}//Kayıtlar.xlsx")
        print(f"""
{Fore.LIGHTWHITE_EX}Lütfen Deponun Yer Aldığı Şehri Seçiniz:
{Fore.CYAN}1 {Fore.RESET}- Tekirdağ
{Fore.CYAN}2 {Fore.RESET}- İzmir
{Fore.CYAN}3 {Fore.RESET}- Çanakkale\n""")
        il_secim = int(input(""))
        if il_secim == 1:
            sayfa_adi = 'Tekirdağ'
        elif il_secim == 2:
            sayfa_adi = 'İzmir'
        elif il_secim == 3:
            sayfa_adi = 'Çanakkale'
        else:
            print(f"{Fore.LIGHTRED_EX}Lütfen Geçerli Bir Sayı Giriniz! (1-3)\n")
            playsound(f"{cwd}//assests//error_sound.wav", block=False)
            return urun_ekle()
        ws = wb[sayfa_adi]
        barkod = str(input(f"\n{Fore.LIGHTWHITE_EX}Lütfen Ürünün Barkod Numarasını Giriniz{Fore.RESET}\n"))
        bos_hucre = 2
        while ws[f'A{bos_hucre}'].value is not None:
            if ws[f'A{bos_hucre}'].value == barkod:
                miktar = ws[f'C{bos_hucre}'].value
                ws[f'C{bos_hucre}'].value = miktar + 1
                print(f"\n{Fore.LIGHTGREEN_EX}Ürün Halihazırda Bulunuyor. Ürünün Miktarı Bir Arttırıldı.\n")
                wb.save(f"{cwd}//Kayıtlar.xlsx")
                break
            else:
                bos_hucre += 1
        while ws[f'A{bos_hucre}'].value is None:
            kenarlik = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                              bottom=Side(style='thin'))
            # Tüm kenarlık aktf sitilini kenarlik değişkenine atar
            ws[f'A{bos_hucre}'].value = barkod
            ws[f'A{bos_hucre}'].alignment = Alignment(horizontal='center')
            ws[f'A{bos_hucre}'].border = kenarlik
            urun_adi = str(input(f"\n{Fore.LIGHTWHITE_EX}Lütfen Ürünün Adını Giriniz\n"))
            ws[f'B{bos_hucre}'].value = urun_adi
            ws[f'B{bos_hucre}'].alignment = Alignment(horizontal='center')
            ws[f'B{bos_hucre}'].border = kenarlik
            urun_miktar = int(input(f"\n{Fore.LIGHTWHITE_EX}Lütfen Ürünün Miktarını Giriniz\n"))
            if urun_miktar >= 0:
                ws[f'C{bos_hucre}'].value = urun_miktar
                ws[f'C{bos_hucre}'].alignment = Alignment(horizontal='center')
                ws[f'C{bos_hucre}'].border = kenarlik
                wb.save(f"{cwd}//Kayıtlar.xlsx")
                break
            else:
                print(f"\n{Fore.LIGHTRED_EX}Lütfen Geçerli Bir Miktar Giriniz!\n")
                playsound(f"{cwd}//assests//error_sound.wav", block=False)
                return urun_ekle()
        tekrar = int(input(f"""{Fore.LIGHTWHITE_EX}
Başka Ürün Eklemek İster Misiniz?
{Fore.CYAN}1 {Fore.RESET}- Evet
{Fore.CYAN}2 {Fore.RESET}- Hayır\n\n"""))
        if tekrar == 1:
            urun_ekle()
        elif tekrar == 2:
            pass
        else:
            print(f"\n{Fore.LIGHTRED_EX}Lütfen Geçerli Bir Sayı Giriniz! (1-3)\n")
            playsound(f"{cwd}//assests//error_sound.wav", block=False)
    except ValueError:
        print(f"\n{Fore.LIGHTRED_EX}Lütfen Bir Sayı Giriniz!\n")
        playsound(f"{cwd}//assests//error_sound.wav", block=False)
    except PermissionError:
        print(f"\n{Fore.LIGHTRED_EX}Excel Dosyası Halihazırda Açık! Lütfen Değişiklik Yapmak İçin Dosyayı Kapatın!")
        playsound(f"{cwd}//assests//error_sound.wav", block=False)
    except FileNotFoundError:
        print(f"\n{Fore.LIGHTRED_EX}Excel Dosyası Bulunamadı. Yeni Bir Tane Oluşturuluyor...\n")
        playsound(f"{cwd}//assests//error_sound.wav", block=False)
        time.sleep(1.13)
        tablo_olustur()
    except Exception as error:
        print(f"{Fore.LIGHTRED_EX}Beklenmeyen Bir Hata Oluştu!\nHata Kodunu Yazdırmak İster Misiniz?\n")
        hata_secim = int(input(f"\n{Fore.CYAN}1{Fore.RESET} - Evet\n{Fore.CYAN}2{Fore.RESET} - Hayır\n"))
        if hata_secim == 1:
            with open("Hatalar.txt", "a", encoding='utf-8') as error_file:
                error_file.write(f"{tarih} Tarihinde Bu Hata Alınmıştır:\n{error}\n\n")
            print(f"{Fore.LIGHTRED_EX}Sistemden Çıkış Yapılıyor...")
            playsound(f"{cwd}//assests//shutdown_sound.wav", block=False)
            time.sleep(1.13)
            exit()
        else:
            print(f"{Fore.LIGHTRED_EX}Sistemden Çıkış Yapılıyor...")
            playsound(f"{cwd}//assests//shutdown_sound.wav", block=False)
            time.sleep(1.13)
            exit()


def urun_bilgi():
    try:
        wb = load_workbook(f"{cwd}//Kayıtlar.xlsx")
        print(f"""
{Fore.LIGHTWHITE_EX}Lütfen Deponun Yer Aldığı Şehri Seçiniz:
{Fore.CYAN}1 {Fore.RESET}- Tekirdağ
{Fore.CYAN}2 {Fore.RESET}- İzmir
{Fore.CYAN}3 {Fore.RESET}- Çanakkale\n""")
        il_secim = int(input(""))
        if il_secim == 1:
            sayfa_adi = 'Tekirdağ'
        elif il_secim == 2:
            sayfa_adi = 'İzmir'
        elif il_secim == 3:
            sayfa_adi = 'Çanakkale'
        else:
            print(f"\n{Fore.LIGHTRED_EX}Lütfen Geçerli Bir Sayı Giriniz! (1-3)\n")
            playsound(f"{cwd}//assests//error_sound.wav", block=False)
            return urun_bilgi()
        ws = wb[sayfa_adi]
        arama_secim = int(input(f"""{Fore.LIGHTWHITE_EX}
Seçiminizi Yapınız:
{Fore.CYAN}1 {Fore.RESET}- Tüm Listeyi Göster
{Fore.CYAN}2 {Fore.RESET}- Barkod Numarsına Ait Ürünü Göster\n\n"""))
        if arama_secim == 1:
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Tüm satırları dolaşır ve içindeki hücre değerlerini alır
                if row[0] != "None" or row[1] != "None" or row[2] != "None":
                    print(f"""
{Fore.LIGHTWHITE_EX}Ürün Bilgileri:
{Fore.LIGHTCYAN_EX}Barkod: {Fore.RESET}{row[0]}
{Fore.LIGHTCYAN_EX}Ürün Adı: {Fore.RESET}{row[1]}
{Fore.LIGHTCYAN_EX}Miktar: {Fore.RESET}{row[2]}\n""")
        elif arama_secim == 2:
            barkod = str(input(f"\n{Fore.LIGHTWHITE_EX}Lütfen Ürünün Barkod Numarasını Giriniz{Fore.RESET}\n"))
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == barkod:
                    print(f"""
{Fore.LIGHTWHITE_EX}Ürün Bilgileri:
{Fore.LIGHTCYAN_EX}Barkod: {Fore.RESET}{row[0]}
{Fore.LIGHTCYAN_EX}Ürün Adı: {Fore.RESET}{row[1]}
{Fore.LIGHTCYAN_EX}Miktar: {Fore.RESET}{row[2]}\n""")
                    break
                else:
                    print(f"\n{Fore.LIGHTRED_EX}Ürün Bulunamadı!\n")
                    playsound(f"{cwd}//assests//error_sound.wav", block=False)
                    return urun_bilgi()
        else:
            print(f"\n{Fore.LIGHTRED_EX}Lütfen Geçerli Bir Sayı Giriniz! (1-2)\n")
            playsound(f"{cwd}//assests//error_sound.wav", block=False)
            # 'block=False' ifadesi sesin bitmesini beklemeden kodun çalışmasına devam etmesini sağlıyor
            return urun_bilgi()
    except ValueError:
        print(f"\n{Fore.LIGHTRED_EX}Lütfen Bir Sayı Giriniz!\n")
        playsound(f"{cwd}//assests//error_sound.wav", block=False)
        return urun_bilgi()
    except PermissionError:
        print(f"\n{Fore.LIGHTRED_EX}Excel Dosyası Halihazırda Açık! Lütfen Değişiklik Yapmak İçin Dosyayı Kapatın!")
        playsound(f"{cwd}//assests//error_sound.wav", block=False)
    except FileNotFoundError:
        print(f"\n{Fore.LIGHTRED_EX}Excel Dosyası Bulunamadı. Yeni Bir Tane Oluşturuluyor...\n")
        playsound(f"{cwd}//assests//error_sound.wav", block=False)
        time.sleep(1.13)
        tablo_olustur()
    except Exception as error:
        print(f"{Fore.LIGHTRED_EX}Beklenmeyen Bir Hata Oluştu!\nHata Kodunu Yazdırmak İster Misiniz?\n")
        hata_secim = int(input(f"\n{Fore.CYAN}1{Fore.RESET} - Evet\n{Fore.CYAN}2{Fore.RESET} - Hayır\n"))
        if hata_secim == 1:
            with open("Hatalar.txt", "a", encoding='utf-8') as error_file:
                error_file.write(f"{tarih} Tarihinde Bu Hata Alınmıştır:\n{error}\n\n")
            print(f"{Fore.LIGHTRED_EX}Sistemden Çıkış Yapılıyor...")
            playsound(f"{cwd}//assests//shutdown_sound.wav", block=False)
            time.sleep(1.13)
            exit()

        else:
            print(f"{Fore.LIGHTRED_EX}Sistemden Çıkış Yapılıyor...")
            playsound(f"{cwd}//assests//shutdown_sound.wav", block=False)
            time.sleep(1.13)
            exit()


def urun_sil():
    try:
        wb = load_workbook(f"{cwd}//Kayıtlar.xlsx")
        print(f"""
{Fore.LIGHTWHITE_EX}Lütfen Deponun Yer Aldığı Şehri Seçiniz:
{Fore.CYAN}1 {Fore.RESET}- Tekirdağ
{Fore.CYAN}2 {Fore.RESET}- İzmir
{Fore.CYAN}3 {Fore.RESET}- Çanakkale\n""")
        il_secim = int(input(""))
        if il_secim == 1:
            sayfa_adi = 'Tekirdağ'
        elif il_secim == 2:
            sayfa_adi = 'İzmir'
        elif il_secim == 3:
            sayfa_adi = 'Çanakkale'
        else:
            print(f"\n{Fore.LIGHTRED_EX}Lütfen Geçerli Bir Sayı Giriniz! (1-3)\n")
            playsound("assests/error_sound.wav", block=False)
            return urun_sil()
        ws = wb[sayfa_adi]
        silme_secim = int(input(f"""
{Fore.LIGHTWHITE_EX}Lütfen Seçiminizi Yapınız:
{Fore.CYAN}1 {Fore.RESET}- Barkod Numarasına Ait Ürünü Sil
{Fore.CYAN}2 {Fore.RESET}- Sayfanın Tamamını Sil\n\n"""))
        if silme_secim == 1:
            barkod = str(input(f"\n{Fore.LIGHTWHITE_EX}Lütfen Ürünün Barkod Numarasını Giriniz{Fore.RESET}\n"))
            urun_bulundu = False
            for row in ws.iter_rows(min_col=1):
                if row[0].value == barkod:
                    ws.delete_rows(row[0].row)
                    urun_bulundu = True
                    print(f"\n{Fore.LIGHTGREEN_EX}Ürün Başarıyla Silindi.\n")
                    wb.save(f"{cwd}//Kayıtlar.xlsx")
                    break
            if not urun_bulundu:
                print(f"\n{Fore.LIGHTRED_EX}Ürün Bulunamadı!\n")
                playsound(f"{cwd}//assests//error_sound.wav", block=False)
                return urun_sil()
        elif silme_secim == 2:
            dogrulama = input(
                f"\n{Fore.LIGHTRED_EX}Tüm Sayfayı Silmek İstediğinizden Emin Misiniz? (E/H){Fore.RESET}\n")
            if dogrulama.upper() == "E" or dogrulama.upper() == "EVET":
                while ws.max_row > 1:
                    ws.delete_rows(2)
                wb.save(f"{cwd}//Kayıtlar.xlsx")
                print(f"{Fore.LIGHTGREEN_EX}Sayfa Başarıyla Silindi.\n")
            else:
                print(f"\n{Fore.LIGHTRED_EX}İşlem İptal Edildi.\n")
        else:
            print(f"\n{Fore.LIGHTRED_EX}Lütfen Geçerli Bir Sayı Giriniz! (1-3)\n")
            playsound("assests/error_sound.wav", block=False)
            return urun_sil()
    except ValueError:
        print(f"\n{Fore.LIGHTRED_EX}Lütfen Bir Sayı Giriniz!\n")
        playsound("assests/error_sound.wav", block=False)
        return urun_sil()
    except PermissionError:
        print(f"\n{Fore.LIGHTRED_EX}Excel Dosyası Halihazırda Açık! Lütfen Değişiklik Yapmak İçin Dosyayı Kapatın!")
        playsound(f"{cwd}//assests//error_sound.wav", block=False)
    except FileNotFoundError:
        print(f"\n{Fore.LIGHTRED_EX}Excel Dosyası Bulunamadı. Yeni Bir Tane Oluşturuluyor...\n")
        playsound(f"{cwd}//assests//error_sound.wav", block=False)
        time.sleep(1.13)
        tablo_olustur()
    except Exception as error:
        print(f"{Fore.LIGHTRED_EX}Beklenmeyen Bir Hata Oluştu!\nHata Kodunu Yazdırmak İster Misiniz?\n")
        hata_secim = int(input(f"\n{Fore.CYAN}1{Fore.RESET} - Evet\n{Fore.CYAN}2{Fore.RESET} - Hayır\n"))
        if hata_secim == 1:
            with open("Hatalar.txt", "a", encoding='utf-8') as error_file:
                error_file.write(f"{tarih} Tarihinde Bu Hata Alınmıştır:\n{error}\n\n")
            print(f"{Fore.LIGHTRED_EX}Sistemden Çıkış Yapılıyor...")
            playsound(f"{cwd}//assests//shutdown_sound.wav", block=False)
            time.sleep(1.13)
            exit()
        else:
            print(f"{Fore.LIGHTRED_EX}Sistemden Çıkış Yapılıyor...")
            playsound(f"{cwd}//assests//shutdown_sound.wav", block=False)
            time.sleep(1.13)
            exit()


def urun_transfer():
    try:
        wb = load_workbook(f"{cwd}//Kayıtlar.xlsx")
        print(f"""
{Fore.LIGHTWHITE_EX}Lütfen Deponun Yer Aldığı Şehri Seçiniz:
{Fore.CYAN}1 {Fore.RESET}- Tekirdağ
{Fore.CYAN}2 {Fore.RESET}- İzmir
{Fore.CYAN}3 {Fore.RESET}- Çanakkale\n""")
        il_secim = int(input(""))
        if il_secim == 1:
            sayfa_adi = 'Tekirdağ'
        elif il_secim == 2:
            sayfa_adi = 'İzmir'
        elif il_secim == 3:
            sayfa_adi = 'Çanakkale'
        else:
            print(f"\n{Fore.LIGHTRED_EX}Lütfen Geçerli Bir Sayı Giriniz! (1-3)\n")
            playsound(f"{cwd}//assests//error_sound.wav", block=False)
            return urun_transfer()
        ws = wb[sayfa_adi]
        barkod = str(input(f"\n{Fore.LIGHTWHITE_EX}Lütfen Ürünün Barkod Numarasını Giriniz{Fore.RESET}\n"))
        for row_data in ws.iter_rows(min_row=2, values_only=True):
            if row_data[0] == barkod:
                print(f"""
{Fore.LIGHTWHITE_EX}Ürün Bilgileri:
{Fore.LIGHTCYAN_EX}Barkod: {Fore.RESET}{row_data[0]}
{Fore.LIGHTCYAN_EX}Ürün Adı: {Fore.RESET}{row_data[1]}
{Fore.LIGHTCYAN_EX}Miktar: {Fore.RESET}{row_data[2]}""")
                print(f"""{Fore.LIGHTWHITE_EX}
Lütfen Seçili Ürünün Transfer Edileceği Depoyu Seçiniz: 
{Fore.CYAN}1 {Fore.RESET}- Tekirdağ
{Fore.CYAN}2 {Fore.RESET}- İzmir
{Fore.CYAN}3 {Fore.RESET}- Çanakkale\n""")
                transfer_il_secim = int(input(""))
                if transfer_il_secim != il_secim:
                    if transfer_il_secim == 1:
                        hedef_sayfa = 'Tekirdağ'
                    elif transfer_il_secim == 2:
                        hedef_sayfa = 'İzmir'
                    elif transfer_il_secim == 3:
                        hedef_sayfa = 'Çanakkale'
                    else:
                        print(f"\n{Fore.LIGHTRED_EX}Lütfen Geçerli Bir Sayı Giriniz! (1-3)\n")
                        playsound(f"{cwd}//assests//error_sound.wav", block=False)
                        return urun_transfer()
                    hedef_ws = wb[hedef_sayfa]
                    bos_hucre = 2
                    while hedef_ws[f'A{bos_hucre}'].value is not None:
                        if hedef_ws[f'A{bos_hucre}'].value == barkod:
                            ws_miktar = row_data[2]
                            hedef_miktar = hedef_ws[f'C{bos_hucre}'].value
                            hedef_ws[f'C{bos_hucre}'].value = ws_miktar + hedef_miktar
                            for row in ws.iter_rows(min_col=1):
                                if row[0].value == barkod:
                                    ws.delete_rows(row[0].row)
                                    break
                            print(f"\n{Fore.LIGHTGREEN_EX}Ürün Halihazırda Bulunuyor. Ürün Miktarı Arttırıldı.")
                            wb.save(f"{cwd}//Kayıtlar.xlsx")
                            break
                        else:
                            bos_hucre += 1
                    else:
                        kenarlik = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                          bottom=Side(style='thin'))
                        hedef_ws[f'A{bos_hucre}'].value = row_data[0]
                        hedef_ws[f'A{bos_hucre}'].alignment = Alignment(horizontal='center')
                        hedef_ws[f'A{bos_hucre}'].border = kenarlik
                        hedef_ws[f'B{bos_hucre}'].value = row_data[1]
                        hedef_ws[f'B{bos_hucre}'].alignment = Alignment(horizontal='center')
                        hedef_ws[f'B{bos_hucre}'].border = kenarlik
                        hedef_ws[f'C{bos_hucre}'].value = row_data[2]
                        hedef_ws[f'C{bos_hucre}'].alignment = Alignment(horizontal='center')
                        hedef_ws[f'C{bos_hucre}'].border = kenarlik
                        for row in ws.iter_rows(min_col=1):
                            if row[0].value == barkod:
                                ws.delete_rows(row[0].row)
                                break
                        print(f"\n{Fore.LIGHTGREEN_EX}Ürününüz Başarıyla Transfer Edilmiştir.\n")
                        wb.save(f"{cwd}//Kayıtlar.xlsx")
                        break
                    break
        else:
            print(f"\n{Fore.LIGHTRED_EX}Ürün Bulunamadı!\n")
            playsound(f"{cwd}//assests//error_sound.wav", block=False)
    except ValueError:
        print(f"\n{Fore.LIGHTRED_EX}Lütfen Bir Sayı Giriniz!\n")
        playsound("assests/error_sound.wav", block=False)
        return urun_transfer()
    except PermissionError:
        print(f"\n{Fore.LIGHTRED_EX}Excel Dosyası Halihazırda Açık! Lütfen Değişiklik Yapmak İçin Dosyayı Kapatın!")
        playsound(f"{cwd}//assests//error_sound.wav", block=False)
    except FileNotFoundError:
        print(f"\n{Fore.LIGHTRED_EX}Excel Dosyası Bulunamadı. Yeni Bir Tane Oluşturuluyor...\n")
        playsound(f"{cwd}//assests//error_sound.wav", block=False)
        time.sleep(1.13)
        tablo_olustur()
    except Exception as error:
        print(f"{Fore.LIGHTRED_EX}Beklenmeyen Bir Hata Oluştu!\nHata Kodunu Yazdırmak İster Misiniz?\n")
        hata_secim = int(input(f"\n{Fore.CYAN}1{Fore.RESET} - Evet\n{Fore.CYAN}2{Fore.RESET} - Hayır\n"))
        if hata_secim == 1:
            with open("Hatalar.txt", "a", encoding='utf-8') as error_file:
                error_file.write(f"{tarih} Tarihinde Bu Hata Alınmıştır:\n{error}\n\n")
            print(f"{Fore.LIGHTRED_EX}Sistemden Çıkış Yapılıyor...")
            playsound(f"{cwd}//assests//shutdown_sound.wav", block=False)
            time.sleep(1.13)
            exit()
        else:
            print(f"{Fore.LIGHTRED_EX}Sistemden Çıkış Yapılıyor...")
            playsound(f"{cwd}//assests//shutdown_sound.wav", block=False)
            time.sleep(1.13)
            exit()


def logo():
    print(f"""{rastgele_renk}
        ____                      ______      __   _          _____ _      __                 _ 
       / __ \\___  ____  ____     /_  __/___ _/ /__(_)___     / ___/(_)____/ /____  ____ ___  (_)
      / / / / _ \\/ __ \\/ __ \\     / / / __ `/ //_/ / __ \\    \\__ \\/ / ___/ __/ _ \\/ __ `__ \\/ / 
     / /_/ /  __/ /_/ / /_/ /    / / / /_/ / ,< / / /_/ /   ___/ / /\\_  / /_/  __/ / / / / / /  
    /_____/\\___/ .___/\\____/    /_/  \\____/_/|_/_/ .___/   /____/_/____/\\__/\\___/_/ /_/ /_/_/              
              /_/                               /_/   
    {Fore.LIGHTWHITE_EX}
    Uygulamasına Hoş Geldin {bilgisayar_adi}!""")

def menu(username):
    logo()
    while True:
        print(f"""
{Fore.LIGHTWHITE_EX}Lütfen Yapmak İstediğiniz İşlemi Seçiniz:
{Fore.CYAN}1 {Fore.RESET}- Ürün Ekle
{Fore.CYAN}2 {Fore.RESET}- Ürün Bilgilerini Göster
{Fore.CYAN}3 {Fore.RESET}- Ürün Bilgilerini Veya Logları Mail Gönder
{Fore.CYAN}4 {Fore.RESET}- Maili Hızlı Erişimlere Ekle
{Fore.CYAN}5 {Fore.RESET}- Ürün Sil
{Fore.CYAN}6 {Fore.RESET}- Ürün Transfer Et
{Fore.LIGHTMAGENTA_EX}7 {Fore.RESET}- Arka Plan Müziğini Oynat
{Fore.LIGHTMAGENTA_EX}8 {Fore.RESET}- Arka Plan Müziğini Durdur
{Fore.LIGHTRED_EX}9 {Fore.RESET}- Çıkış Yap
""")
        try:
            secim = int(input(""))
            if username != "admin" and secim in [1, 5, 6]:
                print(f"\n{Fore.LIGHTRED_EX}Bu İşlemi Yapmaya Yetkiniz Yok!")
                playsound(f"{cwd}//assests//error_sound.wav", block=False)
                continue
            if secim == 1:
                urun_ekle()
            elif secim == 2:
                urun_bilgi()
            elif secim == 3:
                mail_gonder()
            elif secim == 4:
                mail = input("Mail: ")
                hizli_erisim_ekle(mail)
            elif secim == 5:
                urun_sil()
            elif secim == 6:
                urun_transfer()
            elif secim == 7:
                print(f"""{Fore.LIGHTWHITE_EX}
Lütfen İstediğiniz Şarkıyı Seçiniz:
{Fore.CYAN}1{Fore.RESET} - Night Vibes
{Fore.CYAN}2{Fore.RESET} - Lofi Music\n""")
                try:
                    muzik_secim = int(input(""))
                    if muzik_secim == 1:
                        winsound.PlaySound(f"{cwd}//assests//background_music1.wav",
                                           winsound.SND_LOOP + winsound.SND_ASYNC)
                    elif muzik_secim == 2:
                        winsound.PlaySound(f"{cwd}//assests//background_music2.wav",
                                           winsound.SND_LOOP + winsound.SND_ASYNC)
                    else:
                        print(f"\n{Fore.LIGHTRED_EX}Lütfen Geçerli Bir Sayı Giriniz! (1-2)\n")
                        playsound(f"{cwd}//assests//error_sound.wav", block=False)
                except ValueError:
                    print(f"\n{Fore.LIGHTRED_EX}Lütfen Bir Sayı Giriniz!\n")
                    playsound(f"{cwd}//assests//error_sound.wav", block=False)
            elif secim == 8:
                winsound.PlaySound(None, winsound.SND_PURGE)
            elif secim == 9:
                print(f"\n{Fore.LIGHTRED_EX}Hesaptan Çıkılıyor...")
                time.sleep(1.13)
                return giris()
            else:
                print(f"\n{Fore.LIGHTRED_EX}Lütfen Geçerli Bir Sayı Giriniz! (1-4)\n")
                playsound(f"{cwd}//assests//error_sound.wav", block=False)
        except ValueError:
            print(f"\n{Fore.LIGHTRED_EX}Lütfen Bir Sayı Giriniz!\n")
            playsound(f"{cwd}//assests//error_sound.wav", block=False)
        except Exception as error:
            print(f"{Fore.LIGHTRED_EX}Beklenmeyen Bir Hata Oluştu!\nHata Kodunu Yazdırmak İster Misiniz?\n")
            hata_secim = int(input(f"\n{Fore.CYAN}1{Fore.RESET} - Evet\n{Fore.CYAN}2{Fore.RESET} - Hayır\n"))
            if hata_secim == 1:
                with open("Hatalar.txt", "a", encoding='utf-8') as error_file:
                    error_file.write(f"{tarih} Tarihinde Bu Hata Alınmıştır:\n{error}\n\n")
                print(f"{Fore.LIGHTRED_EX}Sistemden Çıkış Yapılıyor...")
                playsound(f"{cwd}//assests//shutdown_sound.wav", block=False)
                time.sleep(1.13)
                exit()
            else:
                print(f"{Fore.LIGHTRED_EX}Sistemden Çıkış Yapılıyor...")
                playsound(f"{cwd}//assests//shutdown_sound.wav", block=False)
                time.sleep(1.13)
                exit()


with open("Loglar.txt", "a", encoding='utf-8') as log_file:
    log_file.write(
        f"{tarih} Tarihinde '{bilgisayar_adi}' Adlı, İp Adresi '{ip_adresi}' Olan Cihaz Sisteme Giriş Yapmıştır\n")
    log_file.close()


def giris():
    try:
        while True:
            ilk_secim = int(input(f"""{Fore.LIGHTWHITE_EX}Lütfen Seçiminizi Yapınız:
{Fore.CYAN}1{Fore.RESET} - Giriş Yap
{Fore.CYAN}2{Fore.RESET} - Kayıt Ol
{Fore.CYAN}3{Fore.RESET} - Şifre Değiştir
{Fore.CYAN}4{Fore.RESET} - Kullanıcı Sil
{Fore.CYAN}5{Fore.RESET} - Kullanıcı Bilgileri
{Fore.LIGHTRED_EX}6{Fore.RESET} - Çıkış Yap
\n"""))
            if ilk_secim == 1:
                while True:
                    username = input(f"{Fore.LIGHTWHITE_EX}Kullanıcı Adı: {Fore.RESET}").lower()
                    password = input(f"{Fore.LIGHTWHITE_EX}Şifre: {Fore.RESET}")
                    password = hashlib.sha512(password.encode('utf-8')).hexdigest()
                    if sifre_giris(username, password):
                        print(f"{Fore.LIGHTGREEN_EX}\nGiriş başarılı.")
                        playsound(f"{cwd}//assests//logon_sound.wav", block=False)
                        time.sleep(1.13)
                        menu(username)
                        break
                    else:
                        print(f"\n{Fore.LIGHTRED_EX}Kullanıcı adı veya şifre hatalı.")
                        playsound(f"{cwd}//assests//error_sound.wav", block=False)
            elif ilk_secim == 2:
                username = input("Kullanıcı Adı: ").lower()
                mail = input("Mail: ")
                password = input("Şifre: ")
                password = hashlib.sha512(password.encode('utf-8')).hexdigest()
                kullanici_ekle(username, mail, password)
                continue
            elif ilk_secim == 3:
                username = input("Kullanıcı Adı: ").lower()
                dogrulama_maili_gonder(username)
            elif ilk_secim == 4:
                while True:
                    username = input("Kullanıcı Adı: ").lower()
                    password = input("Şifre: ")
                    password = hashlib.sha512(password.encode('utf-8')).hexdigest()
                    if kullanici_sil(username, password):
                        break
            elif ilk_secim == 5:
                username = input("Kullanıcı Adı: ").lower()
                sonuc = kullanici_ara(username)
                if not sonuc:
                    print(f"{Fore.LIGHTRED_EX}\nKullanıcı Bulunamadı!\n")
                    playsound(f"{cwd}//assests//error_sound.wav", block=False)
                else:
                    print("")
                    for user in sonuc:
                        print("Kullanıcı Adı: ", user[0])
                        print("E-posta: ", user[1])
                        print("")
            elif ilk_secim == 6:
                print(f"{Fore.LIGHTRED_EX}Çıkılıyor...")
                playsound(f"{cwd}//assests//shutdown_sound.wav", block=False)
                time.sleep(1.13)
                exit()
            else:
                print(f"{Fore.LIGHTRED_EX}\nLütfen Geçerli Bir Sayı Giriniz! (1-6)\n")
                playsound(f"{cwd}//assests//error_sound.wav", block=False)
    except ValueError:
        print(f"{Fore.LIGHTRED_EX}\nLütfen Bir Sayı Giriniz!\n")
        playsound(f"{cwd}//assests//error_sound.wav", block=False)
        return giris()

giris()
